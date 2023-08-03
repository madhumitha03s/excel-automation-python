# import the necessary libraries
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# define the process for a file input
def process_workbook(filename):

    # load the excel workbook and select the sheet
    wb = xl.load_workbook('transactions.xlsx')
    sheet = wb['Sheet1']

    # two ways to pick cells in the sheet

    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)
    # print(sheet.value)

    # perform the operation for each field
    for row in range(2, sheet.max_row + 1):

        # select the required cell and find the percentage
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9

        # write the percentage found into a new cell
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # update the values in the sheet
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # add a bar chart showing the percentage stats
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # save the changes in the same file
    wb.save(filename)
